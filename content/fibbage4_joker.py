import json
from openpyxl import load_workbook


INPUT_FILE = "fibbage_questions.xlsx"
OUTPUT_FILE = "fibbage_questions.lua"

# lua или json
OUTPUT_FORMAT = "lua"

# Добавлять точку в конец последней строки text
ADD_PERIOD = True


def normalize_answer(value):
    value = str(value).strip().lower()

    if value in ("true", "1", "yes", "да", "истина"):
        return True

    if value in ("false", "0", "no", "нет", "ложь"):
        return False

    raise ValueError(f"Не могу распознать answer: {value}")


def add_period_if_needed(text):
    text = str(text).strip()

    if text and text[-1] not in ".!?":
        return text + "."

    return text


def read_questions_from_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    questions = []

    current = {
        "text": [],
        "answer": None
    }

    for row in ws.iter_rows(min_row=1, values_only=True):
        key = row[0]
        value = row[1] if len(row) > 1 else None

        # Пустая строка = конец одного вопроса
        if key is None and value is None:
            if current["text"] and current["answer"] is not None:
                questions.append(current)

            current = {
                "text": [],
                "answer": None
            }
            continue

        if key is None:
            continue

        key = str(key).strip().lower()

        if value is None:
            value = ""

        if key.startswith("text"):
            current["text"].append(str(value).strip())

        elif key == "answer":
            current["answer"] = normalize_answer(value)

    # Добавляем последний вопрос, если после него нет пустой строки
    if current["text"] and current["answer"] is not None:
        questions.append(current)

    if ADD_PERIOD:
        for question in questions:
            if question["text"]:
                question["text"][-1] = add_period_if_needed(question["text"][-1])

    return questions


def save_as_json(questions, output_file):
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(questions, f, ensure_ascii=False, indent=4)


def escape_lua_string(text):
    return text.replace("\\", "\\\\").replace('"', '\\"')


def save_as_lua(questions, output_file):
    lines = []
    lines.append("return {")

    for question in questions:
        lines.append("\t{")
        lines.append("\t\ttext = {")

        for text_line in question["text"]:
            escaped = escape_lua_string(text_line)
            lines.append(f'\t\t\t"{escaped}",')

        lines.append("\t\t},")
        answer = "true" if question["answer"] else "false"
        lines.append(f"\t\tanswer = {answer}")
        lines.append("\t},")

    lines.append("}")

    with open(output_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


questions = read_questions_from_excel(INPUT_FILE)

if OUTPUT_FORMAT == "json":
    save_as_json(questions, OUTPUT_FILE)
else:
    save_as_lua(questions, OUTPUT_FILE)

print("Готово!")
print("Файл создан:", OUTPUT_FILE)
print("Вопросов обработано:", len(questions))
